# coding=utf8
import string
import re

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import openpyxl
import xlrd


def readExcel(excel_path, columns_index):
    excel = xlrd.open_workbook(excel_path)
    sheet = excel.sheets()[0]
    columns = []
    for i in columns_index:
        column = sheet.col_values(colx=i, start_rowx=1, end_rowx=None)
        columns.append(column)
    return columns


def writeExcel(sheet_name, header, data, excel_path):
    workbook = openpyxl.Workbook()
    font = Font(bold=True)
    up = string.ascii_uppercase
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(header)
    for row in data:
        sheet.append(row)
    for i in range(0, sheet.max_column):
        sheet["%s1" % up[i]].font = font
    workbook.save(filename=excel_path)
    workbook.close()